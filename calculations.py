# calculations.py
# IAA休闲游戏成本收益核心计算模块

from __future__ import annotations

import io
import os
import pickle
from datetime import datetime, timedelta
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import curve_fit


# ---------------------------------------------------------------------------
# 常量定义
# ---------------------------------------------------------------------------

# DataFrame 列索引常量，提升可读性
COL_DAY = 0           # 天数索引
COL_UA_COST = 1       # 日均消耗（UA成本）
COL_REVENUE = 2       # 当日收入
COL_GROSS_PROFIT = 3  # 账面毛利
COL_LABOR_COST = 4    # 总用工成本
COL_OTHER_COST = 5    # 其他运营成本
COL_OTHER_TOTAL = 6   # 其他成本合计
COL_TOTAL_COST = 7    # 总成本
COL_NET_PROFIT = 8    # 账面净利
COL_ROI = 9           # 账面ROI
COL_CUM_UA = 10       # 累计消耗
COL_CUM_REV = 11      # 累计收入
COL_CUM_GROSS = 12    # 总毛利
COL_CUM_LABOR = 13    # 累计用工成本
COL_CUM_OTHER = 14    # 累计其他运营成本
COL_CUM_OTHER_TOTAL = 15  # 累计其他成本
COL_CUM_TOTAL_COST = 16   # 累计总成本
COL_CUM_NET = 17      # 总净利
COL_CUM_NET2 = 18     # 总净利2
COL_CPI = 19          # CPI
COL_DNU = 20          # DNU（万人）
COL_DAU = 21          # DAU（万人）
COL_ARPU = 22         # ARPU
COL_CASH_REV_1M = 23  # 滞后1个月累计现金收入
COL_CASH_GAP_1M = 24  # 滞后1个月累计现金缺口
COL_CASH_REV_2M = 25  # 滞后2个月累计现金收入
COL_CASH_GAP_2M = 26  # 滞后2个月累计现金缺口
COL_CUM_PROFIT_FLAG = 27   # 累积利润回本标记
COL_CUM_PROFIT_SUM = 28    # 累积利润回本倒序求和
COL_DYN_PROFIT_FLAG = 29   # 当期利润回本标记
COL_DYN_PROFIT_SUM = 30    # 当期利润回本倒序求和
COL_CASH_1M_FLAG = 31      # 现金流打正标记（滞后1个月）
COL_CASH_1M_SUM = 32       # 现金流打正倒序求和（滞后1个月）
COL_CASH_2M_FLAG = 33      # 现金流打正标记（滞后2个月）
COL_CASH_2M_SUM = 34       # 现金流打正倒序求和（滞后2个月）
COL_DAU_10M_FLAG = 35      # 达成1000万DAU标记
COL_DAU_2M_FLAG = 36       # 达成200万DAU标记
COL_DAU_TARGET_FLAG = 37   # 达成目标DAU标记
COL_ROI_DAILY = 38         # 每日ROI增量
COL_RETENTION = 39         # 留存率

TOTAL_DF_COLS = 40  # DataFrame总列数


# ---------------------------------------------------------------------------
# 辅助函数：曲线拟合
# ---------------------------------------------------------------------------

def fit_roi_curve_advanced(
    known_days: list[int],
    known_values: list[float],
    target_days: int,
    curve_type: str,
) -> list[float]:
    """使用曲线拟合生成完整的ROI或留存率向量。

    Args:
        known_days: 已知数据点的天数索引列表（0-based）。
        known_values: 已知数据点的值列表。
        target_days: 需要生成的总天数。
        curve_type: 曲线类型，"roi" 或 "retention"。

    Returns:
        长度为 target_days 的拟合值列表。
    """
    if not known_days:
        return [0.0] * target_days

    x_data = np.array(known_days, dtype=float)
    y_data = np.array(known_values, dtype=float)

    if curve_type == "roi":
        # ROI 使用幂函数：y = a * x^b + c
        def fit_func(x, a, b, c):
            return a * np.power(x, b) + c

        initial_guess = [y_data[0], 0.5, 0.0]
        bounds = ([0, 0, -np.inf], [np.inf, 2, np.inf])
        max_fev = 10000
    else:
        # 留存率使用衰减幂函数：y = a * (x + 1)^-b + c
        def fit_func(x, a, b, c):
            return a * np.power(x + 1, -b) + c

        initial_guess = [1.0, 0.5, 0.0]
        bounds = ([0, 0, 0], [1.0, np.inf, 1.0])
        max_fev = 5000

    try:
        p_opt, _ = curve_fit(
            fit_func, x_data, y_data,
            p0=initial_guess, bounds=bounds, maxfev=max_fev,
        )
    except Exception as e:
        print(f"[fit_roi_curve_advanced] 拟合失败（{curve_type}）: {e}，使用最后已知值填充")
        return [float(y_data[-1])] * target_days

    all_days = np.arange(target_days, dtype=float)
    fitted = fit_func(all_days, *p_opt)

    # 后处理：确保值在合理范围内
    if curve_type == "roi":
        # ROI 单调不减且非负
        for i in range(1, len(fitted)):
            fitted[i] = max(fitted[i], fitted[i - 1])
        fitted = np.maximum(fitted, 0.0)
    else:
        # 留存率限制在 [0, 1]
        fitted = np.clip(fitted, 0.0, 1.0)

    return fitted.tolist()


# ---------------------------------------------------------------------------
# 辅助函数：曲线向量解析
# ---------------------------------------------------------------------------

def _parse_curve_vector(
    curve_data: dict | None,
    fallback_vector: list[float],
    total_days: int,
    curve_type: str,
) -> np.ndarray:
    """解析单条曲线数据，返回长度为 total_days 的 numpy 数组。

    若 curve_data 为空或无数据点，则返回 fallback_vector 的副本。

    Args:
        curve_data: 曲线数据字典，包含 type 和 points 字段。
        fallback_vector: 全局曲线向量（备用）。
        total_days: 目标天数。
        curve_type: "roi" 或 "retention"。

    Returns:
        长度为 total_days 的 numpy 数组。
    """
    if not curve_data or not curve_data.get("points"):
        return np.array(fallback_vector, dtype=float)

    points = curve_data["points"]
    input_type = curve_data.get("type", "manual")

    if input_type == "excel":
        raw = [p["value"] / 100.0 for p in points]
        if len(raw) >= total_days:
            return np.array(raw[:total_days], dtype=float)
        known_days = list(range(len(raw)))
        fitted = fit_roi_curve_advanced(known_days, raw, total_days, curve_type)
        vec = raw + fitted[len(raw):]
        return np.array(vec, dtype=float)
    else:
        # 手动输入：按天数排序
        pairs = sorted(
            ((p["day"] - 1, p["value"] / 100.0) for p in points),
            key=lambda t: t[0],
        )
        known_days, known_values = zip(*pairs)
        known_days = list(known_days)
        known_values = list(known_values)

        if len(known_values) >= total_days:
            return np.array(known_values[:total_days], dtype=float)
        fitted = fit_roi_curve_advanced(known_days, known_values, total_days, curve_type)
        return np.array(fitted, dtype=float)


def _parse_global_curve(
    data_raw: dict,
    total_days: int,
    curve_type: str,
) -> list[float]:
    """解析全局曲线数据（ROI 或留存率），返回长度为 total_days 的列表。

    Args:
        data_raw: 包含 type 和 points 的原始数据字典。
        total_days: 目标天数。
        curve_type: "roi" 或 "retention"。

    Returns:
        长度为 total_days 的浮点数列表。
    """
    points = data_raw.get("points", [])
    input_type = data_raw.get("type", "manual")

    if input_type == "excel":
        raw = [p["value"] / 100.0 for p in points]
        if len(raw) >= total_days:
            return raw[:total_days]
        known_days = list(range(len(raw)))
        fitted = fit_roi_curve_advanced(known_days, raw, total_days, curve_type)
        return raw + fitted[len(raw):]
    else:
        if not points:
            return [0.0] * total_days
        pairs = sorted(
            ((p["day"] - 1, p["value"] / 100.0) for p in points),
            key=lambda t: t[0],
        )
        known_days, known_values = zip(*pairs)
        known_days = list(known_days)
        known_values = list(known_values)

        if len(known_values) >= total_days:
            return known_values[:total_days]
        return fit_roi_curve_advanced(known_days, known_values, total_days, curve_type)


# ---------------------------------------------------------------------------
# 辅助函数：找到第一个满足条件的天数
# ---------------------------------------------------------------------------

def _first_zero_day(series: pd.Series) -> int:
    """返回 series 中第一个值为 0 的对应天数索引，未找到返回 -1。"""
    idx = series[series == 0].index
    if len(idx) == 0:
        return -1
    return int(idx[0])


def _first_exceed_day(dau_series: pd.Series, threshold: float) -> int:
    """返回 DAU 序列中第一个超过阈值的天数索引，未找到返回 -1。"""
    idx = dau_series[dau_series > threshold].index
    if len(idx) == 0:
        return -1
    return int(idx[0])


# ---------------------------------------------------------------------------
# 主计算函数
# ---------------------------------------------------------------------------

def calculate_metrics(params: dict) -> dict:
    """根据前端传入的参数，计算所有IAA游戏相关的成本收益指标。

    Args:
        params: 包含所有前端输入的字典，结构详见模块文档。

    Returns:
        包含所有计算结果的字典，用于前端展示。
    """
    # ------------------------------------------------------------------
    # 1. 读取基础参数
    # ------------------------------------------------------------------
    project_name: str = params.get("project_name", "default")
    repayment_months: int = params.get("repayment_months", 1)
    repayment_flag: bool = repayment_months > 2
    target_dau: float = params.get("target_dau", 500)

    # ------------------------------------------------------------------
    # 2. 解析投资时间段
    # ------------------------------------------------------------------
    investment_periods_raw: list[dict] = params.get("investment_periods", [])
    if not investment_periods_raw:
        raise ValueError("investment_periods 不能为空")

    investment_periods: list[dict] = []
    total_true_investment_days = 0

    for period in investment_periods_raw:
        start_date = datetime.strptime(period["start"], "%Y-%m-%d")
        end_date = datetime.strptime(period["end"], "%Y-%m-%d")
        period_days = (end_date - start_date).days + 1
        total_true_investment_days += period_days
        investment_periods.append({
            "start": start_date,
            "end": end_date,
            "days": period_days,
            "cost_type": period.get("cost_type", "fixed"),
            "cost_value": period.get("cost_value", 0.0),
            "cost_start": period.get("cost_start", 0.0),
            "cost_end": period.get("cost_end", 0.0),
            "dnu": period.get("dnu", 0.0),
            "team_size": period.get("team_size", 0),
            "labor_cost": period.get("labor_cost", 0.0),
            "other_cost": period.get("other_cost", 0.0),
            "roi_curve_data": period.get("roi_curve_data"),
            "retention_curve_data": period.get("retention_curve_data"),
        })

    # 总天数 = 真实投资天数 + 回款延迟天数（用于现金流计算）
    total_days = total_true_investment_days + min(repayment_months, 2) * 30

    # ------------------------------------------------------------------
    # 3. 解析全局 ROI 和留存率曲线
    # ------------------------------------------------------------------
    roi_vector = _parse_global_curve(
        params.get("roi_data", {"type": "manual", "points": []}),
        total_days, "roi",
    )
    retention_vector = _parse_global_curve(
        params.get("retention_data", {"type": "manual", "points": []}),
        total_days, "retention",
    )

    # ------------------------------------------------------------------
    # 4. 构建每日数据 DataFrame
    # ------------------------------------------------------------------
    df = _build_daily_dataframe(
        investment_periods, roi_vector, retention_vector,
        total_days, total_true_investment_days, target_dau,
    )

    # ------------------------------------------------------------------
    # 5. 处理 n > 2 个月回款的扩展现金流
    # ------------------------------------------------------------------
    df_extra = _build_extra_cashflow(df, repayment_months, repayment_flag, total_days)

    # ------------------------------------------------------------------
    # 6. 计算关键指标
    # ------------------------------------------------------------------
    key_metrics = _compute_key_metrics(
        df, df_extra, repayment_flag, repayment_months, target_dau,
    )

    # ------------------------------------------------------------------
    # 7. 按季度聚合，生成图表和表格数据
    # ------------------------------------------------------------------
    charts, quarterly_table_data = _aggregate_quarterly(
        df, df_extra, investment_periods, total_true_investment_days, repayment_flag,
    )

    # ------------------------------------------------------------------
    # 8. 持久化 DataFrame 到本地文件
    # ------------------------------------------------------------------
    safe_project_name = _safe_filename(project_name)
    _save_dataframe(safe_project_name, df, df_extra, repayment_months, repayment_flag)

    # ------------------------------------------------------------------
    # 9. 构建并返回结果
    # ------------------------------------------------------------------
    results: dict[str, Any] = {
        "key_metrics": key_metrics,
        "charts": charts,
        "quarterly_table_data": quarterly_table_data,
        "data_file_saved": True,
        "project_name": safe_project_name,
    }
    return results


# ---------------------------------------------------------------------------
# 内部函数：构建每日 DataFrame
# ---------------------------------------------------------------------------

def _build_daily_dataframe(
    investment_periods: list[dict],
    roi_vector: list[float],
    retention_vector: list[float],
    total_days: int,
    total_true_investment_days: int,
    target_dau: float,
) -> pd.DataFrame:
    """构建包含所有每日计算数据的 DataFrame。"""
    df = pd.DataFrame(np.zeros((total_days, TOTAL_DF_COLS)))
    df.iloc[:, COL_DAY] = np.arange(total_days, dtype=float)

    # --- 填充各时间段的基础数据，并解析各时间段专属曲线 ---
    period_roi_vecs: list[np.ndarray] = []
    period_ret_vecs: list[np.ndarray] = []
    current_day = 0

    for period in investment_periods:
        period_days = period["days"]
        sl = slice(current_day, current_day + period_days)

        # UA 日消耗
        if period["cost_type"] == "fixed":
            df.iloc[sl, COL_UA_COST] = period["cost_value"]
        else:
            df.iloc[sl, COL_UA_COST] = np.linspace(
                period["cost_start"], period["cost_end"], period_days
            )

        # 用工成本 & 其他运营成本
        df.iloc[sl, COL_LABOR_COST] = period["team_size"] * period["labor_cost"]
        df.iloc[sl, COL_OTHER_COST] = period["other_cost"]

        # DNU
        df.iloc[sl, COL_DNU] = period["dnu"]

        # 解析该时间段的 ROI / 留存率曲线
        period_roi_vecs.append(
            _parse_curve_vector(period.get("roi_curve_data"), roi_vector, total_days, "roi")
        )
        period_ret_vecs.append(
            _parse_curve_vector(period.get("retention_curve_data"), retention_vector, total_days, "retention")
        )

        current_day += period_days

    # 全局 ROI / 留存率（用于 Excel 导出展示）
    df.iloc[:, COL_ROI] = roi_vector
    df.iloc[:, COL_ROI_DAILY] = np.diff(roi_vector, prepend=0.0)
    df.iloc[0, COL_ROI_DAILY] = roi_vector[0]
    df.iloc[:, COL_RETENTION] = retention_vector

    # --- 卷积计算每日收入 ---
    ua_array = df.iloc[:, COL_UA_COST].to_numpy(dtype=float)
    daily_revenue = _convolve_revenue(investment_periods, ua_array, period_roi_vecs, total_days)
    df.iloc[:, COL_REVENUE] = daily_revenue

    # --- 卷积计算每日 DAU ---
    dnu_array = df.iloc[:, COL_DNU].to_numpy(dtype=float)
    daily_dau = _convolve_dau(investment_periods, dnu_array, period_ret_vecs, total_days)
    df.iloc[:, COL_DAU] = daily_dau

    # --- 每日衍生指标（向量化） ---
    ua = df.iloc[:, COL_UA_COST].to_numpy(dtype=float)
    rev = df.iloc[:, COL_REVENUE].to_numpy(dtype=float)
    labor = df.iloc[:, COL_LABOR_COST].to_numpy(dtype=float)
    other = df.iloc[:, COL_OTHER_COST].to_numpy(dtype=float)
    dnu_col = df.iloc[:, COL_DNU].to_numpy(dtype=float)
    dau_col = df.iloc[:, COL_DAU].to_numpy(dtype=float)

    other_total = labor + other
    total_cost = ua + other_total
    gross_profit = rev - ua
    net_profit = gross_profit - other_total

    cum_ua = np.cumsum(ua)
    cum_rev = np.cumsum(rev)
    cum_gross = np.cumsum(gross_profit)
    cum_labor = np.cumsum(labor)
    cum_other = np.cumsum(other)
    cum_other_total = np.cumsum(other_total)
    cum_total_cost = cum_other_total + cum_ua
    cum_net = cum_gross - cum_other_total
    cum_net2 = cum_rev - cum_total_cost

    # 避免除零
    with np.errstate(divide="ignore", invalid="ignore"):
        cpi = np.where(dnu_col != 0, ua / dnu_col, 0.0)
        arpu = np.where(dau_col != 0, rev / dau_col, 0.0)

    # 现金流（滞后 1 / 2 个月）
    cum_rev_series = pd.Series(cum_rev)
    cash_rev_1m = cum_rev_series.shift(30).fillna(0.0).to_numpy()
    cash_rev_2m = cum_rev_series.shift(60).fillna(0.0).to_numpy()
    cash_gap_1m = cash_rev_1m - cum_total_cost
    cash_gap_2m = cash_rev_2m - cum_total_cost

    # 回本 / 打正标记（0=已达成，1=未达成）
    cum_profit_flag = (cum_net <= 0).astype(int)
    dyn_profit_flag = (net_profit <= 0).astype(int)
    cash_1m_flag = (cash_gap_1m <= 0).astype(int)
    cash_2m_flag = (cash_gap_2m <= 0).astype(int)

    # DAU 目标标记
    dau_10m_flag = (dau_col < 1000).astype(int)
    dau_2m_flag = (dau_col < 200).astype(int)
    dau_target_flag = (dau_col < target_dau).astype(int)

    # 倒序累计求和（用于找到"最后一天仍未达成"的天数）
    def _reverse_cumsum(arr: np.ndarray) -> np.ndarray:
        return arr[::-1].cumsum()[::-1]

    # 写回 DataFrame
    assignments = {
        COL_GROSS_PROFIT: gross_profit,
        COL_OTHER_TOTAL: other_total,
        COL_TOTAL_COST: total_cost,
        COL_NET_PROFIT: net_profit,
        COL_CUM_UA: cum_ua,
        COL_CUM_REV: cum_rev,
        COL_CUM_GROSS: cum_gross,
        COL_CUM_LABOR: cum_labor,
        COL_CUM_OTHER: cum_other,
        COL_CUM_OTHER_TOTAL: cum_other_total,
        COL_CUM_TOTAL_COST: cum_total_cost,
        COL_CUM_NET: cum_net,
        COL_CUM_NET2: cum_net2,
        COL_CPI: cpi,
        COL_ARPU: arpu,
        COL_CASH_REV_1M: cash_rev_1m,
        COL_CASH_GAP_1M: cash_gap_1m,
        COL_CASH_REV_2M: cash_rev_2m,
        COL_CASH_GAP_2M: cash_gap_2m,
        COL_CUM_PROFIT_FLAG: cum_profit_flag,
        COL_CUM_PROFIT_SUM: _reverse_cumsum(cum_profit_flag),
        COL_DYN_PROFIT_FLAG: dyn_profit_flag,
        COL_DYN_PROFIT_SUM: _reverse_cumsum(dyn_profit_flag),
        COL_CASH_1M_FLAG: cash_1m_flag,
        COL_CASH_1M_SUM: _reverse_cumsum(cash_1m_flag),
        COL_CASH_2M_FLAG: cash_2m_flag,
        COL_CASH_2M_SUM: _reverse_cumsum(cash_2m_flag),
        COL_DAU_10M_FLAG: dau_10m_flag,
        COL_DAU_2M_FLAG: dau_2m_flag,
        COL_DAU_TARGET_FLAG: dau_target_flag,
    }
    for col_idx, values in assignments.items():
        df.iloc[:, col_idx] = values

    return df


def _convolve_revenue(
    investment_periods: list[dict],
    ua_array: np.ndarray,
    period_roi_vecs: list[np.ndarray],
    total_days: int,
) -> np.ndarray:
    """按时间段卷积计算每日收入。"""
    daily_revenue = np.zeros(total_days)
    current_day = 0
    for idx, period in enumerate(investment_periods):
        period_days = period["days"]
        ua_period = np.zeros(total_days)
        ua_period[current_day:current_day + period_days] = ua_array[current_day:current_day + period_days]

        p_roi_vec = period_roi_vecs[idx]
        p_roi_diff = np.diff(p_roi_vec, prepend=0.0)
        p_roi_diff[0] = p_roi_vec[0]

        daily_revenue += np.convolve(ua_period, p_roi_diff)[:total_days]
        current_day += period_days
    return daily_revenue


def _convolve_dau(
    investment_periods: list[dict],
    dnu_array: np.ndarray,
    period_ret_vecs: list[np.ndarray],
    total_days: int,
) -> np.ndarray:
    """按时间段卷积计算每日 DAU。"""
    dau = np.zeros(total_days)
    current_day = 0
    for idx, period in enumerate(investment_periods):
        period_days = period["days"]
        dnu_period = np.zeros(total_days)
        dnu_period[current_day:current_day + period_days] = dnu_array[current_day:current_day + period_days]

        dau += np.convolve(dnu_period, period_ret_vecs[idx])[:total_days]
        current_day += period_days
    return dau


def _build_extra_cashflow(
    df: pd.DataFrame,
    repayment_months: int,
    repayment_flag: bool,
    total_days: int,
) -> pd.DataFrame:
    """构建 n > 2 个月回款的扩展现金流 DataFrame（4列）。"""
    df_extra = pd.DataFrame(np.zeros((total_days, 4)))
    if not repayment_flag:
        return df_extra

    cum_rev = df.iloc[:, COL_CUM_REV].to_numpy(dtype=float)
    cum_total_cost = df.iloc[:, COL_CUM_TOTAL_COST].to_numpy(dtype=float)

    cash_rev_nm = pd.Series(cum_rev).shift(repayment_months * 30).fillna(0.0).to_numpy()
    cash_gap_nm = cash_rev_nm - cum_total_cost
    cash_flag_nm = (cash_gap_nm <= 0).astype(int)
    cash_sum_nm = cash_flag_nm[::-1].cumsum()[::-1]

    df_extra.iloc[:, 0] = cash_rev_nm
    df_extra.iloc[:, 1] = cash_gap_nm
    df_extra.iloc[:, 2] = cash_flag_nm
    df_extra.iloc[:, 3] = cash_sum_nm
    return df_extra


# ---------------------------------------------------------------------------
# 内部函数：计算关键指标
# ---------------------------------------------------------------------------

def _compute_key_metrics(
    df: pd.DataFrame,
    df_extra: pd.DataFrame,
    repayment_flag: bool,
    repayment_months: int,
    target_dau: float,
) -> dict:
    """从 DataFrame 中提取所有关键指标。"""
    cash_gap_1m = df.iloc[:, COL_CASH_GAP_1M]
    cash_gap_2m = df.iloc[:, COL_CASH_GAP_2M]

    max_cash_demand_1m = float(abs(cash_gap_1m.min())) if cash_gap_1m.min() < 0 else 0.0
    max_cash_demand_2m = float(abs(cash_gap_2m.min())) if cash_gap_2m.min() < 0 else 0.0

    # 使用 COL_DAY 列作为索引（天数从0开始）
    day_series = df.iloc[:, COL_DAY].astype(int)

    dynamic_profit_breakeven_day = _find_breakeven_day(df.iloc[:, COL_DYN_PROFIT_SUM], day_series)
    cumulative_profit_breakeven_day = _find_breakeven_day(df.iloc[:, COL_CUM_PROFIT_SUM], day_series)
    cumulative_cash_flow_1m_breakeven_day = _find_breakeven_day(df.iloc[:, COL_CASH_1M_SUM], day_series)
    cumulative_cash_flow_2m_breakeven_day = _find_breakeven_day(df.iloc[:, COL_CASH_2M_SUM], day_series)

    dau_col = df.iloc[:, COL_DAU]
    day_to_10m_dau = _find_first_exceed(dau_col, 1000, day_series)
    day_to_2m_dau = _find_first_exceed(dau_col, 200, day_series)
    day_to_target_dau = _find_first_exceed(dau_col, target_dau, day_series)

    key_metrics: dict[str, Any] = {
        "max_cash_demand_1m": round(max_cash_demand_1m, 2),
        "max_cash_demand_2m": round(max_cash_demand_2m, 2),
        "dynamic_profit_breakeven_day": dynamic_profit_breakeven_day,
        "cumulative_profit_breakeven_day": cumulative_profit_breakeven_day,
        "cumulative_cash_flow_1m_breakeven_day": cumulative_cash_flow_1m_breakeven_day,
        "cumulative_cash_flow_2m_breakeven_day": cumulative_cash_flow_2m_breakeven_day,
        "day_to_10m_dau": day_to_10m_dau,
        "day_to_2m_dau": day_to_2m_dau,
        "day_to_target_dau": day_to_target_dau,
        "target_dau": target_dau,
        "repayment_months": repayment_months,
    }

    if repayment_flag:
        cash_gap_nm = df_extra.iloc[:, 1]
        max_cash_demand_nm = float(abs(cash_gap_nm.min())) if cash_gap_nm.min() < 0 else 0.0
        nm_day_series = pd.Series(range(len(df_extra)))
        cumulative_cash_flow_nm_breakeven_day = _find_breakeven_day(
            df_extra.iloc[:, 3], nm_day_series
        )
        key_metrics["max_cash_demand_nm"] = round(max_cash_demand_nm, 2)
        key_metrics["cumulative_cash_flow_nm_breakeven_day"] = cumulative_cash_flow_nm_breakeven_day

    return key_metrics


def _find_breakeven_day(sum_series: pd.Series, day_series: pd.Series) -> int:
    """找到倒序求和列中第一个为 0 的对应天数，未找到返回 -1。"""
    mask = sum_series.to_numpy() == 0
    idx = np.argmax(mask)
    if not mask[idx]:
        return -1
    return int(day_series.iloc[idx])


def _find_first_exceed(dau_series: pd.Series, threshold: float, day_series: pd.Series) -> int:
    """找到 DAU 序列中第一个超过阈值的对应天数，未找到返回 -1。"""
    mask = dau_series.to_numpy() > threshold
    idx = np.argmax(mask)
    if not mask[idx]:
        return -1
    return int(day_series.iloc[idx])


# ---------------------------------------------------------------------------
# 内部函数：季度聚合
# ---------------------------------------------------------------------------

def _aggregate_quarterly(
    df: pd.DataFrame,
    df_extra: pd.DataFrame,
    investment_periods: list[dict],
    total_true_investment_days: int,
    repayment_flag: bool,
) -> tuple[dict, list[dict]]:
    """按季度聚合数据，返回图表数据和季度表格数据。"""
    base_date = investment_periods[0]["start"]
    df = df.copy()
    df["date"] = pd.to_datetime([
        base_date + timedelta(days=int(d)) for d in df.iloc[:, COL_DAY]
    ])
    df["quarter"] = df["date"].dt.to_period("Q").astype(str).str.replace("20", "", regex=False)

    # 只取真实投资天数范围内的数据
    df_display = df.iloc[:total_true_investment_days].copy()
    quarters = df_display["quarter"].unique().tolist()

    dau_labels, dau_data = [], []
    fin_labels, fin_income, fin_cost, fin_cum_profit = [], [], [], []
    quarterly_table_data = []

    prev = {
        "revenue": 0.0, "cost": 0.0, "ua_cost": 0.0,
        "personnel_cost": 0.0, "other_cost": 0.0, "profit": 0.0,
        "cash_1m": 0.0, "cash_2m": 0.0, "cash_nm": 0.0,
    }

    for quarter in quarters:
        q_df = df_display[df_display["quarter"] == quarter]
        if q_df.empty:
            continue

        end_idx = q_df.index[-1]
        end_date = q_df["date"].iloc[-1]

        cum_days = int(df_display.loc[end_idx, COL_DAY]) + 1
        cum_rev = round(float(df_display.loc[end_idx, COL_CUM_REV]), 2)
        cum_total_cost = -round(float(df_display.loc[end_idx, COL_CUM_TOTAL_COST]), 2)
        cum_ua_cost = -round(float(df_display.loc[end_idx, COL_CUM_UA]), 2)
        cum_personnel = -round(float(df_display.loc[end_idx, COL_CUM_LABOR]), 2)
        cum_other = -round(float(df_display.loc[end_idx, COL_CUM_OTHER]), 2)

        raw_1m = df_display.loc[end_idx, COL_CASH_GAP_1M]
        raw_2m = df_display.loc[end_idx, COL_CASH_GAP_2M]
        cum_cash_1m = round(float(raw_1m) if not pd.isna(raw_1m) else 0.0, 2)
        cum_cash_2m = round(float(raw_2m) if not pd.isna(raw_2m) else 0.0, 2)

        cum_cash_nm = 0.0
        if repayment_flag:
            raw_nm = df_extra.loc[end_idx, 1] if end_idx < len(df_extra) else 0.0
            cum_cash_nm = round(float(raw_nm) if not pd.isna(raw_nm) else 0.0, 2)

        cur_rev = round(cum_rev - prev["revenue"], 2)
        cur_cost = round(cum_total_cost - prev["cost"], 2)
        cur_ua = round(cum_ua_cost - prev["ua_cost"], 2)
        cur_personnel = round(cum_personnel - prev["personnel_cost"], 2)
        cur_other = round(cum_other - prev["other_cost"], 2)
        cur_profit = round(cur_rev + cur_cost, 2)
        cum_profit = round(prev["profit"] + cur_profit, 2)

        cur_demand_1m = max(0.0, -round(cum_cash_1m - prev["cash_1m"], 2))
        cur_demand_2m = max(0.0, -round(cum_cash_2m - prev["cash_2m"], 2))

        dau = round(float(df_display.loc[end_idx, COL_DAU]), 2)

        # 图表数据
        dau_labels.append(quarter)
        dau_data.append(dau)
        fin_labels.append(quarter)
        fin_income.append(cur_rev)
        fin_cost.append(cur_cost)
        fin_cum_profit.append(cum_profit)

        # 表格数据
        row: dict[str, Any] = {
            "quarter": quarter,
            "end_date": end_date.strftime("%Y/%m/%d"),
            "cumulative_days": cum_days,
            "cumulative_revenue": cum_rev,
            "cumulative_cost": cum_total_cost,
            "ua_cost": cum_ua_cost,
            "personnel_cost": cum_personnel,
            "other_cost": cum_other,
            "cumulative_profit": cum_profit,
            "cumulative_cash_flow_1m": cum_cash_1m,
            "cumulative_cash_flow_2m": cum_cash_2m,
            "current_revenue": cur_rev,
            "current_cost": cur_cost,
            "current_ua_cost": cur_ua,
            "current_personnel_cost": cur_personnel,
            "current_other_cost": cur_other,
            "current_profit": cur_profit,
            "current_cash_demand_1m": round(cur_demand_1m, 2),
            "current_cash_demand_2m": round(cur_demand_2m, 2),
            "dau": dau,
        }

        if repayment_flag:
            cur_demand_nm = max(0.0, -round(cum_cash_nm - prev["cash_nm"], 2))
            row["cumulative_cash_flow_nm"] = cum_cash_nm
            row["current_cash_demand_nm"] = round(cur_demand_nm, 2)
            prev["cash_nm"] = cum_cash_nm

        quarterly_table_data.append(row)

        # 更新前一季度累计值
        prev.update({
            "revenue": cum_rev,
            "cost": cum_total_cost,
            "ua_cost": cum_ua_cost,
            "personnel_cost": cum_personnel,
            "other_cost": cum_other,
            "profit": cum_profit,
            "cash_1m": cum_cash_1m,
            "cash_2m": cum_cash_2m,
        })

    charts = {
        "dau_quarterly": {"labels": dau_labels, "data": dau_data},
        "finance_quarterly": {
            "labels": fin_labels,
            "income": fin_income,
            "cost": fin_cost,
            "cumulative_profit": fin_cum_profit,
        },
    }
    return charts, quarterly_table_data


# ---------------------------------------------------------------------------
# 内部函数：文件持久化
# ---------------------------------------------------------------------------

def _safe_filename(project_name: str) -> str:
    """将项目名称转换为安全的文件名。"""
    if not project_name:
        return "default"
    return project_name.replace("/", "_").replace("\\", "_").replace(":", "_")


def _save_dataframe(
    safe_project_name: str,
    df: pd.DataFrame,
    df_extra: pd.DataFrame | None,
    repayment_months: int,
    repayment_flag: bool,
) -> None:
    """将 DataFrame 序列化保存到本地 pickle 文件。"""
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(data_dir, exist_ok=True)

    data_file_path = os.path.join(data_dir, f"{safe_project_name}_data.pkl")
    save_data = {
        "df": df,
        "df_extra": df_extra if repayment_flag else None,
        "repayment_months": repayment_months,
        "repayment_flag": repayment_flag,
    }
    with open(data_file_path, "wb") as f:
        pickle.dump(save_data, f)


# ---------------------------------------------------------------------------
# Excel 导出函数
# ---------------------------------------------------------------------------

def export_daily_excel(project_name: str) -> bytes:
    """读取保存的 DataFrame 文件，转换为 Excel 文件返回。

    Args:
        project_name: 项目名称，用于定位保存的数据文件。

    Returns:
        Excel 文件的二进制数据，用于导出下载。

    Raises:
        FileNotFoundError: 数据文件不存在时抛出。
    """
    safe_project_name = _safe_filename(project_name)
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    data_file_path = os.path.join(data_dir, f"{safe_project_name}_data.pkl")

    if not os.path.exists(data_file_path):
        raise FileNotFoundError(f"数据文件不存在: {data_file_path}")

    with open(data_file_path, "rb") as f:
        loaded_data = pickle.load(f)

    df: pd.DataFrame = loaded_data["df"]
    df_extra: pd.DataFrame | None = loaded_data.get("df_extra")
    repayment_months: int = loaded_data.get("repayment_months", 1)
    repayment_flag: bool = loaded_data.get("repayment_flag", False)

    wb = _build_excel_workbook(df, df_extra, repayment_months, repayment_flag)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_excel_workbook(
    df: pd.DataFrame,
    df_extra: pd.DataFrame | None,
    repayment_months: int,
    repayment_flag: bool,
) -> Workbook:
    """构建 Excel 工作簿。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "每日数据"

    # 样式定义
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    extra_cols = 6 if repayment_flag else 0
    total_columns = TOTAL_DF_COLS + extra_cols

    _write_excel_headers(ws, repayment_months, repayment_flag)
    _write_excel_data(ws, df, df_extra, repayment_flag)
    _apply_excel_styles(ws, header_font, center_align, thin_border, total_columns, len(df))

    return wb


def _write_excel_headers(ws, repayment_months: int, repayment_flag: bool) -> None:
    """写入 Excel 表头（第1-3行）。"""
    # ---------- 第1行：大分类表头 ----------
    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="每日")
    ws.cell(row=1, column=10, value="累计")
    ws.cell(row=1, column=20, value="用户量")
    ws.cell(row=1, column=24, value="现金流")

    target_col = 30 if repayment_flag else 28
    ws.cell(row=1, column=target_col, value="目标达成周期计算")

    ws.merge_cells("B1:I1")
    ws.merge_cells("J1:S1")
    ws.merge_cells("T1:W1")

    if repayment_flag:
        ws.merge_cells("X1:AC1")
        ws.merge_cells("AD1:AR1")
    else:
        ws.merge_cells("X1:AA1")
        ws.merge_cells("AB1:AN1")

    # ---------- 第2行：子分类表头 ----------
    row2 = [""] * 9 + [""] * 10 + [""] * 4 + ["滞后1个月", "", "滞后2个月", ""]
    if repayment_flag:
        row2 += [f"滞后{repayment_months}个月", ""]
    row2 += ["累积利润回本", "", "当期利润回本", "",
             "累计现金流打正（滞后1个月）", "",
             "累计现金流打正（滞后2个月）", ""]
    if repayment_flag:
        row2 += [f"累计现金流打正（滞后{repayment_months}个月）", ""]
    row2 += ["1000万DAU", "200万DAU", "目标DAU", "", ""]

    for col, val in enumerate(row2, 1):
        ws.cell(row=2, column=col, value=val)

    ws.merge_cells("X2:Y2")
    ws.merge_cells("Z2:AA2")

    if repayment_flag:
        for rng in ["AB2:AC2", "AD2:AE2", "AF2:AG2", "AH2:AI2",
                    "AJ2:AK2", "AL2:AM2", "AN2:AO2", "AP2:AQ2"]:
            ws.merge_cells(rng)
    else:
        for rng in ["AB2:AC2", "AD2:AE2", "AF2:AG2", "AH2:AI2",
                    "AJ2:AK2"]:
            ws.merge_cells(rng)

    # ---------- 第3行：具体列名 ----------
    row3 = [
        "days", "预期日均消耗", "当日收入", "账面毛利", "总用工成本", "其他运营成本",
        "其他成本合计", "总成本", "账面净利",
        "账面ROI", "累计消耗", "累计收入", "总毛利", "累计用工成本", "累计其他运营成本",
        "累计其他成本", "累计总成本", "总净利", "总净利2",
        "CPI", "DNU (万)", "DAU (万)", "ARPU",
        "累计现金收入", "累计现金缺口", "累计现金收入", "累计现金缺口",
    ]
    if repayment_flag:
        row3 += ["累计现金收入", "累计现金缺口"]
    row3 += [
        "正累计收益=0，负数=1", "倒序求和", "正当期收益=0，负数=1", "倒序求和",
        "现金流为正=0，为负=1", "倒序求和", "现金流为正=0，为负=1", "倒序求和",
    ]
    if repayment_flag:
        row3 += ["现金流为正=0，为负=1", "倒序求和"]
    row3 += [
        "1000万以上DAU=0，以下=1",
        "200万以上DAU=0，以下=1",
        "xx万以上DAU=0，以下=1",
        "ROI of the day", "Retention",
    ]

    for col, val in enumerate(row3, 1):
        ws.cell(row=3, column=col, value=val)


def _write_excel_data(
    ws,
    df: pd.DataFrame,
    df_extra: pd.DataFrame | None,
    repayment_flag: bool,
) -> None:
    """将 DataFrame 数据写入 Excel（从第4行开始）。"""
    def _safe_round(val, ndigits: int) -> float:
        return round(float(val), ndigits) if not pd.isna(val) else 0.0

    for i in range(len(df)):
        row_num = i + 4
        row_data = [
            int(df.iloc[i, COL_DAY]),
            _safe_round(df.iloc[i, COL_UA_COST], 2),
            _safe_round(df.iloc[i, COL_REVENUE], 2),
            _safe_round(df.iloc[i, COL_GROSS_PROFIT], 2),
            _safe_round(df.iloc[i, COL_LABOR_COST], 2),
            _safe_round(df.iloc[i, COL_OTHER_COST], 2),
            _safe_round(df.iloc[i, COL_OTHER_TOTAL], 2),
            _safe_round(df.iloc[i, COL_TOTAL_COST], 2),
            _safe_round(df.iloc[i, COL_NET_PROFIT], 2),
            _safe_round(df.iloc[i, COL_ROI], 4),
            _safe_round(df.iloc[i, COL_CUM_UA], 2),
            _safe_round(df.iloc[i, COL_CUM_REV], 2),
            _safe_round(df.iloc[i, COL_CUM_GROSS], 2),
            _safe_round(df.iloc[i, COL_CUM_LABOR], 2),
            _safe_round(df.iloc[i, COL_CUM_OTHER], 2),
            _safe_round(df.iloc[i, COL_CUM_OTHER_TOTAL], 2),
            _safe_round(df.iloc[i, COL_CUM_TOTAL_COST], 2),
            _safe_round(df.iloc[i, COL_CUM_NET], 2),
            _safe_round(df.iloc[i, COL_CUM_NET2], 2),
            _safe_round(df.iloc[i, COL_CPI], 4),
            _safe_round(df.iloc[i, COL_DNU], 2),
            _safe_round(df.iloc[i, COL_DAU], 2),
            _safe_round(df.iloc[i, COL_ARPU], 4),
            _safe_round(df.iloc[i, COL_CASH_REV_1M], 2),
            _safe_round(df.iloc[i, COL_CASH_GAP_1M], 2),
            _safe_round(df.iloc[i, COL_CASH_REV_2M], 2),
            _safe_round(df.iloc[i, COL_CASH_GAP_2M], 2),
        ]

        if repayment_flag and df_extra is not None:
            row_data += [
                _safe_round(df_extra.iloc[i, 0], 2),
                _safe_round(df_extra.iloc[i, 1], 2),
            ]

        row_data += [
            int(df.iloc[i, COL_CUM_PROFIT_FLAG]),
            int(df.iloc[i, COL_CUM_PROFIT_SUM]),
            int(df.iloc[i, COL_DYN_PROFIT_FLAG]),
            int(df.iloc[i, COL_DYN_PROFIT_SUM]),
            int(df.iloc[i, COL_CASH_1M_FLAG]),
            int(df.iloc[i, COL_CASH_1M_SUM]),
            int(df.iloc[i, COL_CASH_2M_FLAG]),
            int(df.iloc[i, COL_CASH_2M_SUM]),
        ]

        if repayment_flag and df_extra is not None:
            row_data += [
                int(df_extra.iloc[i, 2]),
                int(df_extra.iloc[i, 3]),
            ]

        row_data += [
            int(df.iloc[i, COL_DAU_10M_FLAG]),
            int(df.iloc[i, COL_DAU_2M_FLAG]),
            int(df.iloc[i, COL_DAU_TARGET_FLAG]),
            _safe_round(df.iloc[i, COL_ROI_DAILY], 4),
            _safe_round(df.iloc[i, COL_RETENTION], 4),
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)


def _apply_excel_styles(
    ws,
    header_font: Font,
    center_align: Alignment,
    thin_border: Border,
    total_columns: int,
    data_rows: int,
) -> None:
    """应用 Excel 样式（表头加粗居中、边框、列宽）。"""
    # 表头样式
    for row in range(1, 4):
        for col in range(1, total_columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    # 数据行边框
    for row in range(4, data_rows + 4):
        for col in range(1, total_columns + 1):
            ws.cell(row=row, column=col).border = thin_border

    # 自适应列宽（采样末尾100行，避免全量遍历）
    total_rows = data_rows + 4
    sample_start = max(1, total_rows - 100)
    for col in range(1, total_columns + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(sample_start, total_rows)
        )
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, 20), 8)